import streamlit as st
import pandas as pd
import datetime
import os
import openpyxl
from docxtpl import DocxTemplate

def enter_data():

    name = user_name
    nachname = user_nachname
    kw = user_kw 

    if name and nachname:
            year = user_jahr
            
            course = berufe
            date_from = user_date_from
            date_to = user_date_to

            # Course info
            # registration_status = reg_status_var.get()
            duration_monday = zeit_montag
            duration_tuesday = zeit_dienstag
            duration_wednesday = zeit_mittwoch
            duration_thursday = zeit_donnerstag
            duration_friday = zeit_freitag

            content_monday = user_content_montag
            content_tuesday = user_content_dienstag
            content_wednesday = user_content_mittwoch
            content_thursday = user_content_donnerstag
            content_friday = user_content_freitag

            excel_template = "excel-template.xlsx"
            filepath = f"Exports_Excel\Arbeitsbericht_KW_{kw}_{year}_{name}_{nachname}_.xlsx"

            if not os.path.exists(filepath):
                        file = openpyxl.load_workbook(excel_template)
                        sheet = file.active
                        #----------------------------------------------------------------
                        sheet['A2'].value = name
                        sheet['B2'].value = nachname
                        sheet['C2'].value = kw
                        sheet['D2'].value = year
                        sheet['E2'].value = course
                        sheet['F2'].value = date_from
                        sheet['G2'].value = date_to
                        #----------------------------------------------------------------
                        sheet['A5'].value = duration_monday
                        sheet['B5'].value = duration_tuesday
                        sheet['C5'].value = duration_wednesday
                        sheet['D5'].value = duration_thursday
                        sheet['E5'].value = duration_friday
                        #----------------------------------------------------------------
                        sheet['B7'].value = content_monday
                        sheet['B8'].value = content_tuesday
                        sheet['B9'].value = content_wednesday
                        sheet['B10'].value = content_thursday
                        sheet['B11'].value = content_friday
                    #----------------------------------------------------------------
                        
                        filepath = f"Exports_Excel\Arbeitsbericht_KW_{kw}_{year}_{name}_{nachname}_.xlsx"
                        file.save(filepath)
                        options_form.success('The data was saved successfully and the file was created successfully.', icon="⚠️")
            
            if os.path.exists(filepath):
                        file = openpyxl.load_workbook(filepath)
                        sheet = file.active
                        #----------------------------------------------------------------
                        sheet['A2'].value = name
                        sheet['B2'].value = nachname
                        sheet['C2'].value = kw
                        sheet['E2'].value = year
                        sheet['D2'].value = course
                        sheet['F2'].value = date_from
                        sheet['G2'].value = date_to
                        #----------------------------------------------------------------
                        sheet['A5'].value = duration_monday
                        sheet['B5'].value = duration_tuesday
                        sheet['C5'].value = duration_wednesday
                        sheet['D5'].value = duration_thursday
                        sheet['E5'].value = duration_friday
                        #----------------------------------------------------------------
                        sheet['B7'].value = content_monday
                        sheet['B8'].value = content_tuesday
                        sheet['B9'].value = content_wednesday
                        sheet['B10'].value = content_thursday
                        sheet['B11'].value = content_friday
                    #----------------------------------------------------------------
                        
                        filepath = f"Exports_Excel\Arbeitsbericht_KW_{kw}_{year}_{name}_{nachname}_.xlsx"
                        file.save(filepath)
                        #tkinter.messagebox.showinfo(title="Success", message="The data was saved successfully and the file was created successfully.")
                        options_form.success('Report has been saved successfully into a Word file.', icon="⚠️")


                        
            else:
                options_form.warning('Name, Surname, KW and Year are required.', icon="⚠️")
                pass
    else:
            
        options_form.warning('Name, Surname, KW and Year are required.', icon="⚠️")
        pass      


def excel_to_word():
        
    name = user_name
    nachname = user_nachname
    kw = user_kw 

    
    year = user_jahr
    
    course = berufe
    date_from = user_date_from
    date_to = user_date_to
    duration_monday = zeit_montag
    duration_tuesday = zeit_dienstag
    duration_wednesday = zeit_mittwoch
    duration_thursday = zeit_donnerstag
    duration_friday = zeit_freitag

    content_monday = user_content_montag
    content_tuesday = user_content_dienstag
    content_wednesday = user_content_mittwoch
    content_thursday = user_content_donnerstag
    content_friday = user_content_freitag

    filepath = f"Exports_Excel\Arbeitsbericht_KW_{kw}_{year}_{name}_{nachname}_.xlsx"
        
    if name and nachname and kw and year:
        if os.path.exists(filepath):
            file = openpyxl.load_workbook(filepath)
            sheet = file.active


            name = sheet['A2'].value
            nachname = sheet['B2'].value
            kw = sheet['C2'].value
            year = sheet['E2'].value
            
            course = sheet['D2'].value
            date_from = sheet['F2'].value
            date_to = sheet['G2'].value

            # Course info
            # registration_status = reg_status_var.get()
            duration_monday = sheet['A5'].value
            duration_tuesday = sheet['B5'].value
            duration_wednesday = sheet['C5'].value
            duration_thursday = sheet['D5'].value
            duration_friday = sheet['E5'].value

            content_monday = sheet['B7'].value
            content_tuesday = sheet['B8'].value
            content_wednesday = sheet['B9'].value
            content_thursday = sheet['B10'].value
            content_friday = sheet['B11'].value
            # Generate docs
            doc = DocxTemplate("example.docx")

            doc.render({    "name": name,
                            "surname": nachname,
                            "kw": kw,
                            "year": year,
                            "course": course,
                            "date_from": date_from,
                            "date_to": date_to,
                            "duration_monday": duration_monday,
                            "duration_tuesday": duration_tuesday,
                            "duration_wednesday": duration_wednesday,
                            "duration_thursday": duration_thursday,
                            "duration_friday": duration_friday,
                            "content_monday": content_monday,
                            "content_tuesday": content_tuesday,
                            "content_wednesday": content_wednesday,
                            "content_thursday": content_thursday,
                            "content_friday": content_friday,
                            })
                
            doc_name = "Exports_Word\Arbeitsbericht_KW_" + kw + "_" + year + "_" + name + "_" + nachname + "_.docx"
                
            doc.save(doc_name)
            options_form.warning('Report has been saved successfully into a Word file.', icon="⚠️")
    
        else:
            options_form.warning('File not found. Please correct the inputs and try again.', icon="⚠️")
    
    else:
        options_form.warning('Name, Surname, KW and Year are required.', icon="⚠️")


pagr_bg_img = """"
<style>
[data-testid="stAppViewContainer"]{
background-image: url("https://images.unsplash.com/photo-1456324504439-367cee3b3c32?q=80&w=2070&auto=format&fit=crop&ixlib=rb-4.0.3&ixid=M3wxMjA3fDB8MHxwaG90by1wYWdlfHx8fGVufDB8fHx8fA%3D%3D");
background-size: cover;
}
</style>
"""

st.markdown(pagr_bg_img, unsafe_allow_html=True)

st.title("Wochenbericht APP - LDS")


st.header("Working File ...")


st.sidebar.title("Hier ausfühlen ...")

st.sidebar.header("User Information")
options_form = st.sidebar.form("user_information")
user_kw = options_form.text_input("Kalenderwoche")
user_name = options_form.text_input("Name")
user_nachname = options_form.text_input("Nachame")
user_jahr = options_form.selectbox(
    'Jahr',
    ('', '2024', '2025', '2026', '2027', '2028', '2029', '2030', '2031', '2032', '2033', '2034', '2035', '2036', '2037', '2038', '2039', '2040', '2041', '2042', '2043', '2044', '2045', '2046', '2047', '2048', '2049', '2050', '2051', '2052', '2053', '2054', '2055', '2056', '2057', '2058', '2059', '2060', '2061', '2062', '2063', '2064', '2065', '2066', '2067', '2068', '2069', '2070', '2071', '2072', '2073', '2074', '2075', '2076', '2077', '2078', '2079', '2080', '2081', '2082', '2083', '2084', '2085', '2086', '2087', '2088', '2089', '2090', '2091', '2092', '2093', '2094', '2095', '2096', '2097', '2098', '2099'))


berufe = options_form.selectbox(
    'Berufe',
    ('', 'Fachinformatiker für Systemintegration', 'Fachinformatiker für Anwendungsentwicklung'))
user_date_from = options_form.date_input("Von (Tag)", value=None)
user_date_to = options_form.date_input("Bis (Tag)", value=None)
user_weekday = options_form.selectbox(
    'Wochentag',
    ('', 'Montag', 'Dienstag', 'Mittwoch', 'Donnerstag', 'Freitag'))
user_content_montag = options_form.text_area("Beschreibung Montag")
zeit_montag = options_form.time_input('Stunden Montag', value=None)
user_content_dienstag = options_form.text_area("Beschreibung Dienstag")
zeit_dienstag = options_form.time_input('Stunden Dienstag', value=None)
user_content_mittwoch = options_form.text_area("Beschreibung Mittwoch")
zeit_mittwoch = options_form.time_input('Stunden Mittwoch', value=None)
user_content_donnerstag = options_form.text_area("Beschreibung Donnerstag")
zeit_donnerstag = options_form.time_input('Stunden Donnerstag', value=None)
user_content_freitag = options_form.text_area("Beschreibung Freitag")
zeit_freitag = options_form.time_input('Stunden Freitag', value=None)

add_data = options_form.form_submit_button(label="Speichern Excel", on_click=enter_data)
add_data = options_form.form_submit_button(label="Speichern Word", on_click=excel_to_word)

